import os
from openpyxl import Workbook,load_workbook

# trans_num,total_num
def write_toxl(name,in_kg:dict,diff:dict,nb, tnb, SN):
    path='/Users/jiaqili/desktop/嘉禾/六月_结果.xlsx'
    jud = os.path.exists(path)
    if jud:
        # write to exist file
        wb= load_workbook(path)
        sName=wb.sheetnames
        if SN in sName:
            ws=wb[SN]
        elif SN not in sName:
            ws = wb.create_sheet(SN, len(sName)+1)
    else:
        # create new file
        wb=Workbook()
        ws=wb.create_sheet(SN,0)

    tableTitle=['员工','产品','盒数','折合盒数','上纸实际用量',	'中纸实际用量','下纸实际用量',	'合计',
                '出纸率','上纸浪费情况','中纸浪费情况',	'下纸浪费情况']

    cell_range = ws['A1':'L1']
    for i in range(len(cell_range[0])):
        if cell_range[0][i].value!=tableTitle[i]:
            cell_range[0][i].value=tableTitle[i]

    for key in in_kg.keys():
        # print(key)
        new_line = []
        new_line.append(name)
        new_line.append(key)
        new_line.append(tnb[key])
        new_line.append(nb[key])
        new_line.extend(in_kg[key])
        s=sum(in_kg[key])
        new_line.append(s)
        new_line.append(round(nb[key]/s*1000,2))
        new_line.extend(diff[key])
        ws.append(new_line)
    wb.save(path)

def write_eachRow(SN,name,each,x,y):
    path = '/Users/jiaqili/desktop/嘉禾/六月_每行.xlsx'
    jud = os.path.exists(path)
    if jud:
        # write to exist file
        wb = load_workbook(path)
        sName = wb.sheetnames
        if SN in sName:
            ws = wb[SN]
        elif SN not in sName:
            ws = wb.create_sheet(SN, len(sName) + 1)
    else:
        # create new file
        wb = Workbook()
        ws = wb.create_sheet(SN, 0)

    tableTitle = ['员工名字','盒数','折合盒数','上纸实际用量', '中纸实际用量', '下纸实际用量','合计']

    cell_range = ws['A1':'G1']
    for i in range(len(cell_range[0])):
        if cell_range[0][i].value != tableTitle[i]:
            cell_range[0][i].value = tableTitle[i]
    new_line=[]
    for i in range(len(each)):
        new_line.append(name)
        new_line.append(y[i])
        new_line.append(x[i])
        new_line.extend(each[i])
        new_line.append(sum(each[i]))
        ws.append(new_line)
        new_line.clear()
    wb.save(path)